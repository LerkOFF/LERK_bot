from openpyxl import load_workbook, Workbook
from config import LOG_FILE_PATH
from datetime import datetime


def log_user_action(action, user):
    try:
        try:
            workbook = load_workbook(LOG_FILE_PATH)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Timestamp", "User", "Action"])

        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        sheet.append([timestamp, str(user), action])

        workbook.save(LOG_FILE_PATH)
    except Exception as e:
        print(f"Ошибка при логировании действия: {e}")
